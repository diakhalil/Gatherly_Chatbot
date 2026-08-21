"""Build an .xlsx planning workbook from SQL context + RAG ops guides."""

from __future__ import annotations

import os
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import re

EXPORT_DIR = Path(__file__).resolve().parents[4] / "exports"
AGENT_BASE_URL = os.getenv("AGENT_API_URL", "http://127.0.0.1:8001")

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_WRAP = Alignment(wrap_text=True, vertical="top")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(bottom=_THIN)


def _style_header(ws, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP
    ws.freeze_panes = "A2"


def _auto_width(ws, min_width: int = 12, max_width: int = 65):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        longest = min_width
        for cell in col_cells:
            if cell.value:
                longest = max(longest, min(len(str(cell.value)), max_width))
            cell.alignment = _WRAP
            cell.border = _BORDER
        ws.column_dimensions[col_letter].width = longest + 2


def _add_event_sheet(wb: Workbook, ctx: dict):
    ws = wb.create_sheet("Event")
    ws.append(["Field", "Value"])
    _style_header(ws, 2)
    event = ctx.get("event") or {}
    for key, label in [
        ("event_id", "Event ID"),
        ("title", "Title"),
        ("type", "Type"),
        ("starts_at", "Starts"),
        ("ends_at", "Ends"),
        ("guest_count", "Guests"),
        ("required_hosts", "Hosts required"),
        ("assigned_hosts", "Hosts assigned"),
        ("status", "Status"),
        ("location", "Location"),
    ]:
        ws.append([label, event.get(key)])

    venue = ctx.get("venue") or {}
    if venue:
        ws.append([])
        ws.append(["— Venue —", ""])
        ws[ws.max_row][0].font = Font(bold=True)
        for key, label in [
            ("name", "Name"),
            ("city", "City"),
            ("capacity", "Capacity"),
            ("indoor_outdoor", "Setting"),
            ("wheelchair_accessible", "Accessible"),
            ("parking_available", "Parking"),
        ]:
            ws.append([label, venue.get(key)])

    clothing = ctx.get("clothing") or {}
    if clothing:
        ws.append([])
        ws.append(["— Dress Code —", ""])
        ws[ws.max_row][0].font = Font(bold=True)
        ws.append(["Label", clothing.get("label")])
        ws.append(["Description", clothing.get("description")])

    _auto_width(ws)


def _add_team_sheet(wb: Workbook, ctx: dict):
    ws = wb.create_sheet("Team")
    ws.append(["Name", "Role", "Needs ride", "Request dress"])
    _style_header(ws, 4)
    for host in ctx.get("hosts") or []:
        ws.append([
            host.get("name"),
            host.get("role"),
            "Yes" if host.get("needs_ride") else "No",
            "Yes" if host.get("request_dress") else "No",
        ])
    _auto_width(ws)


def _safe_sheet_title(title: str, fallback: str = "Guide") -> str:
    clean = re.sub(r"[\\/*?:\[\]]", " ", str(title or "").strip())
    clean = " ".join(clean.split())
    return (clean or fallback)[:31]


def _add_agent_guide_sheet(wb: Workbook, title: str, rows: list[dict]):
    ws = wb.create_sheet(_safe_sheet_title(title))
    ws.append(["Topic", "Detail"])
    _style_header(ws, 2)

    for row in rows:
        detail = str(row.get("detail") or "").strip()
        if not detail:
            continue
        ws.append([
            str(row.get("topic") or "").strip(),
            detail,
        ])

    _auto_width(ws, max_width=80)


def _add_guide_sheet(wb: Workbook, title: str, guides: list[dict]):
    ws = wb.create_sheet(_safe_sheet_title(title))
    ws.append(["Topic", "Detail"])
    _style_header(ws, 2)

    for guide in guides:
        answer = str(guide.get("answer") or "").strip()
        if not answer or "don't have enough information" in answer.lower():
            continue

        for line in answer.split("\n"):
            line = line.strip()
            if not line:
                continue
            clean = line.lstrip("-•* ").strip()
            clean = re.sub(r"\[T\d+\]", "", clean).strip()
            if not clean:
                continue

            if ":" in clean and len(clean.split(":", 1)[0]) < 40:
                topic_part, detail_part = clean.split(":", 1)
                ws.append([topic_part.strip(), detail_part.strip()])
            else:
                ws.append(["", clean])

        sources = guide.get("text_sources") or []
        if sources:
            names = sorted({
                str(s.get("file_name") or "").strip()
                for s in sources
                if s.get("file_name")
            })
            if names:
                ws.append([])
                ws.append(["Sources", ", ".join(names)])

    _auto_width(ws, max_width=80)


def _sheet_title_from_guide(guide: dict, index: int) -> str:
    topics = guide.get("topics") or []
    topic = str(topics[0]).strip() if topics else f"Guide {index}"
    return topic.replace("_", " ").title()



def build_workbook(
    *,
    context: dict[str, Any],
    ops_guides: list[dict],
    sheets: list[str],
    event_id: int | None,
    guide_sections: list[dict] | None = None,
) -> dict[str, Any]:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    clean_sheets = [s.strip().lower() for s in (sheets or []) if s.strip()]
    if not clean_sheets:
        clean_sheets = ["event", "team"]

    include_event = "event" in clean_sheets
    include_team = "team" in clean_sheets or include_event

    wb = Workbook()
    wb.remove(wb.active)

    included = []

    if include_event:
        _add_event_sheet(wb, context)
        included.append("Event")

    if include_team:
        _add_team_sheet(wb, context)
        included.append("Team")

    if guide_sections:
        for section in guide_sections:
            title = str(section.get("title") or "Guide").strip()
            rows = section.get("rows") or []
            if not rows:
                continue
            _add_agent_guide_sheet(wb, title, rows)
            included.append(_safe_sheet_title(title))
    else:
        guide_index = 1
        for guide in ops_guides:
            answer = str(guide.get("answer") or "").strip()
            if not answer or "don't have enough information" in answer.lower():
                continue
            title = _sheet_title_from_guide(guide, guide_index)
            _add_guide_sheet(wb, title, [guide])
            included.append(title)
            guide_index += 1

    if not wb.sheetnames:
        _add_event_sheet(wb, context)
        included.append("Event")

    eid = event_id or context.get("event", {}).get("event_id") or "pack"
    filename = f"planning_pack_event_{eid}.xlsx"
    filepath = EXPORT_DIR / filename
    wb.save(str(filepath))

    download_url = f"{AGENT_BASE_URL}/exports/{filename}"

    return {
        "path": str(filepath),
        "filename": filename,
        "download_url": download_url,
        "sheets_included": included,
    }

